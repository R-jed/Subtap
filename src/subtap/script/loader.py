"""Script file format loader: txt, srt, md, docx, xlsx."""

from __future__ import annotations

import re
from pathlib import Path


class UnsupportedFormatError(Exception):
    """Raised when script file format is not supported."""


def load_script(path: Path) -> str:
    """Load script file and return raw text content.

    Args:
        path: Path to script file.

    Returns:
        Raw text content.

    Raises:
        UnsupportedFormatError: If file format is not supported.
        FileNotFoundError: If file does not exist.
    """
    if not path.exists():
        raise FileNotFoundError(f"文稿文件不存在：{path}")

    suffix = path.suffix.lower()
    if suffix == ".txt":
        return _load_txt(path)
    elif suffix == ".srt":
        return _load_srt(path)
    elif suffix == ".md":
        return _load_md(path)
    elif suffix == ".docx":
        return _load_docx(path)
    elif suffix == ".xlsx":
        return _load_xlsx(path)
    else:
        raise UnsupportedFormatError(f"不支持的文稿格式：{suffix}，请转为 .txt 后重试")


def _load_txt(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def _load_srt(path: Path) -> str:
    """Extract text from SRT file, removing sequence numbers and timestamps."""
    text = path.read_text(encoding="utf-8")
    lines = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        # 跳过序号行（纯数字）
        if re.match(r"^\d+$", line):
            continue
        # 跳过时间轴行
        if "-->" in line:
            continue
        lines.append(line)
    return "\n".join(lines)


def _load_md(path: Path) -> str:
    """Strip Markdown formatting, keep plain text."""
    text = path.read_text(encoding="utf-8")
    lines = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        # 去除标题标记
        line = re.sub(r"^#+\s*", "", line)
        # 去除粗体/斜体
        line = re.sub(r"\*{1,3}(.+?)\*{1,3}", r"\1", line)
        # 去除链接，保留文本
        line = re.sub(r"\[(.+?)\]\(.+?\)", r"\1", line)
        # 去除列表标记
        line = re.sub(r"^[-*+]\s+", "", line)
        # 去除引用标记
        line = re.sub(r"^>\s*", "", line)
        if line:
            lines.append(line)
    return "\n".join(lines)


def _load_docx(path: Path) -> str:
    """Extract paragraph text from docx file."""
    from docx import Document

    doc = Document(str(path))
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    return "\n".join(paragraphs)


def _load_xlsx(path: Path) -> str:
    """Read first column of first sheet as text lines."""
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    lines = []
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        value = row[0]
        if value is not None:
            text = str(value).strip()
            if text:
                lines.append(text)
    wb.close()
    return "\n".join(lines)
